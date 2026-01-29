"""
Serviço para geração de relatório de averbações de seguro.

Gera relatório no formato Excel com dados de processos que tiveram DI registrada
no mês solicitado. Funciona para qualquer categoria (BND, ALH, VDM, etc.).

Estrutura do relatório (baseado na planilha de exemplo):
- País de Origem
- Porto Origem
- País Origem
- Cidade de Destino
- Data do BL
- Tipo de transporte
- Mercadoria
- Nome Navio
- Custo USD
- Frete USD
- Despesas (cálculo: 10% de Custo + Frete)
- Lucros (cálculo: 10% de Custo + Frete + Despesas)
- Impostos da DI USD (soma de todos os impostos)
- DI
- OBS (processo_referencia)
"""

import logging
import json
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta
from pathlib import Path

logger = logging.getLogger(__name__)

# Mapeamento de códigos de impostos
CODIGOS_IMPOSTOS = {
    '0086': 'II',  # IMPOSTO DE IMPORTAÇÃO
    '86': 'II',
    '1038': 'IPI',
    '38': 'IPI',
    '5602': 'PIS',  # PIS/PASEP IMPORTAÇÃO
    '602': 'PIS',
    '5629': 'COFINS',  # COFINS IMPORTAÇÃO
    '629': 'COFINS',
    '7811': 'TAXA_SISCOMEX',  # TAXA DE UTILIZAÇÃO DO SISCOMEX
    '5529': 'ANTIDUMPING',
    '529': 'ANTIDUMPING',
}

# Mapeamento de códigos de país (ISO 2 letras)
PAISES = {
    'CN': 'CHINA',
    'IT': 'ITALIA',
    'US': 'ESTADOS UNIDOS',
    'DE': 'ALEMANHA',
    # Adicionar mais conforme necessário
}

# Mapeamento de códigos de via de transporte
TIPOS_TRANSPORTE = {
    '1': 'NAVIO',
    '2': 'AEREO',
    '3': 'RODOVIARIO',
    '4': 'FERROVIARIO',
}


class RelatorioAverbacoesService:
    """Serviço para geração de relatório de averbações"""
    
    def __init__(self):
        """Inicializa o serviço"""
        self.downloads_dir = Path('downloads')
        self.downloads_dir.mkdir(exist_ok=True)
    
    def gerar_relatorio_averbacoes(
        self,
        mes: str,
        ano: Optional[int] = None,
        categoria: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Gera relatório de averbações para processos com DI registrada no mês.
        
        Args:
            mes: Mês no formato MM (ex: "06") ou YYYY-MM (ex: "2025-06")
            ano: Ano (opcional, se não fornecido no mes)
            categoria: Categoria do processo (opcional, ex: "BND", "ALH", "VDM")
                      Se None, busca todas as categorias
        
        Returns:
            Dict com sucesso, caminho_arquivo, total_processos, erros
        """
        try:
            # Parsear mês/ano
            if '-' in mes:
                # Formato YYYY-MM
                partes = mes.split('-')
                ano = int(partes[0])
                mes_num = int(partes[1])
            else:
                # Formato MM
                mes_num = int(mes)
                if ano is None:
                    ano = datetime.now().year
            
            logger.info(f'📊 Gerando relatório de averbações para {mes_num:02d}/{ano} (categoria: {categoria or "TODAS"})')
            
            # 1. Buscar processos com DI registrada no mês
            processos = self._buscar_processos_com_di_no_mes(mes_num, ano, categoria)
            
            if not processos:
                return {
                    'sucesso': False,
                    'erro': 'NENHUM_PROCESSO',
                    'mensagem': f'Nenhum processo encontrado com DI registrada em {mes_num:02d}/{ano}'
                }
            
            logger.info(f'✅ Encontrados {len(processos)} processos com DI registrada')
            
            # 2. Para cada processo, buscar dados completos da DI via API
            dados_relatorio = []
            erros = []
            
            for idx, processo in enumerate(processos, 1):
                try:
                    logger.info(f'📄 Processando {idx}/{len(processos)}: {processo.get("processo_referencia")} - DI {processo.get("numero_di")}')
                    
                    # Buscar dados completos da DI
                    dados_linha = self._extrair_dados_processo(processo)
                    
                    if dados_linha:
                        dados_relatorio.append(dados_linha)
                    else:
                        erros.append(f'{processo.get("processo_referencia")}: Não foi possível extrair dados da DI')
                        
                except Exception as e:
                    logger.error(f'❌ Erro ao processar {processo.get("processo_referencia")}: {e}', exc_info=True)
                    erros.append(f'{processo.get("processo_referencia")}: {str(e)}')
            
            if not dados_relatorio:
                return {
                    'sucesso': False,
                    'erro': 'NENHUM_DADO',
                    'mensagem': 'Nenhum dado foi extraído dos processos',
                    'erros': erros
                }
            
            # 3. Gerar Excel
            caminho_arquivo = self._gerar_excel(dados_relatorio, mes_num, ano, categoria)
            
            return {
                'sucesso': True,
                'caminho_arquivo': str(caminho_arquivo),
                'total_processos': len(dados_relatorio),
                'total_erros': len(erros),
                'erros': erros if erros else None
            }
            
        except Exception as e:
            logger.error(f'❌ Erro ao gerar relatório: {e}', exc_info=True)
            return {
                'sucesso': False,
                'erro': 'ERRO_INTERNO',
                'mensagem': f'Erro ao gerar relatório: {str(e)}'
            }
    
    def _buscar_processos_com_di_no_mes(
        self,
        mes: int,
        ano: int,
        categoria: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Busca processos com DI registrada no mês especificado.
        
        Args:
            mes: Mês (1-12)
            ano: Ano
            categoria: Categoria do processo (opcional)
        
        Returns:
            Lista de processos com DI registrada no mês
        """
        try:
            from utils.sql_server_adapter import get_sql_adapter
            from services.db_policy_service import get_primary_database
            
            sql_adapter = get_sql_adapter()
            database_para_usar = get_primary_database()
            
            # Data inicial e final do mês
            data_inicio = datetime(ano, mes, 1)
            if mes == 12:
                data_fim = datetime(ano + 1, 1, 1) - timedelta(seconds=1)
            else:
                data_fim = datetime(ano, mes + 1, 1) - timedelta(seconds=1)
            
            # ✅ CORREÇÃO: Usar mesma estrutura do relatório FOB que funciona
            # Começar de PROCESSO_IMPORTACAO e fazer JOIN com tabelas de DI
            query = '''
                SELECT DISTINCT
                    p.numero_processo,
                    p.id_processo_importacao,
                    p.id_importacao,
                    p.numero_di,
                    diDesp.dataHoraDesembaraco,
                    diDesp.dataHoraRegistro,
                    ddg.numeroDi,
                    ddg.dataHoraSituacaoDi
                FROM {database}.dbo.PROCESSO_IMPORTACAO p WITH (NOLOCK)
                INNER JOIN Serpro.dbo.Hi_Historico_Di diH WITH (NOLOCK)
                    ON p.id_importacao = diH.idImportacao
                INNER JOIN Serpro.dbo.Di_Root_Declaracao_Importacao diRoot WITH (NOLOCK)
                    ON diH.diId = diRoot.dadosDiId
                INNER JOIN Serpro.dbo.Di_Dados_Gerais ddg WITH (NOLOCK)
                    ON diRoot.dadosGeraisId = ddg.dadosGeraisId
                LEFT JOIN Serpro.dbo.Di_Dados_Despacho diDesp WITH (NOLOCK)
                    ON diRoot.dadosDespachoId = diDesp.dadosDespachoId
                WHERE 1=1
                    AND (
                        -- Filtrar por dataHoraDesembaraco se existir
                        (diDesp.dataHoraDesembaraco IS NOT NULL 
                         AND YEAR(diDesp.dataHoraDesembaraco) = ? 
                         AND MONTH(diDesp.dataHoraDesembaraco) = ?)
                        OR
                        -- Senão usar dataHoraSituacaoDi
                        (diDesp.dataHoraDesembaraco IS NULL 
                         AND ddg.dataHoraSituacaoDi IS NOT NULL
                         AND YEAR(ddg.dataHoraSituacaoDi) = ? 
                         AND MONTH(ddg.dataHoraSituacaoDi) = ?)
                        OR
                        -- Fallback: usar dataHoraRegistro se existir
                        (diDesp.dataHoraDesembaraco IS NULL 
                         AND ddg.dataHoraSituacaoDi IS NULL
                         AND diDesp.dataHoraRegistro IS NOT NULL
                         AND YEAR(diDesp.dataHoraRegistro) = ? 
                         AND MONTH(diDesp.dataHoraRegistro) = ?)
                    )
            '''
            
            params = [ano, mes, ano, mes, ano, mes]
            
            # Se categoria foi especificada, filtrar
            if categoria:
                categoria_upper = categoria.upper().strip()
                query += ' AND p.numero_processo LIKE ?'
                params.append(f'{categoria_upper}.%')
            
            query += ' ORDER BY p.numero_processo'
            
            # ✅ POLÍTICA CENTRAL: Usar banco primário (mAIke_assistente)
            # As tabelas de DI estão no banco Serpro (cross-database query)
            query_formatada = query.format(database=database_para_usar)
            logger.info(f'🔍 Executando query para buscar DIs em {mes:02d}/{ano} (categoria: {categoria or "TODAS"}, banco: {database_para_usar})')
            result = sql_adapter.execute_query(query_formatada, database_para_usar, params, notificar_erro=False)
            
            if not result.get('success'):
                logger.error(f'❌ Erro na query SQL: {result.get("error")}')
                return []
            
            if result.get('data'):
                processos = []
                for row in result['data']:
                    processos.append({
                        'processo_referencia': row.get('numero_processo'),
                        'id_processo_importacao': row.get('id_processo_importacao'),
                        'id_importacao': row.get('id_importacao'),
                        'numero_di': row.get('numeroDi') or row.get('numero_di'),
                        'data_registro': row.get('dataHoraDesembaraco') or row.get('dataHoraRegistro') or row.get('dataHoraSituacaoDi')
                    })
                logger.info(f'✅ Query retornou {len(processos)} processos')
                return processos
            
            logger.warning(f'⚠️ Query retornou sucesso mas sem dados')
            return []
            
        except Exception as e:
            logger.error(f'❌ Erro ao buscar processos: {e}', exc_info=True)
            return []
    
    def _buscar_di_sql_server(self, numero_di: str, numero_di_normalizado: str) -> Optional[Dict[str, Any]]:
        """
        Busca dados da DI no SQL Server (prioridade sobre API bilhetada).
        
        Busca nome do navio da tabela Di_Transporte (campo nomeVeiculo).
        
        Args:
            numero_di: Número da DI original
            numero_di_normalizado: Número da DI normalizado (sem / e -)
        
        Returns:
            Dict com dados da DI no formato similar à API ou None se não encontrado
        """
        try:
            from utils.sql_server_adapter import get_sql_adapter
            
            sql_adapter = get_sql_adapter()
            
            # Query para buscar dados da DI incluindo nome do navio
            # Baseada na query documentada em MAPEAMENTO_SQL_SERVER.md
            # ✅ CORREÇÃO (26/01/2026): Di_Dados_Embarque não existe no Serpro. Usar Di_Transporte.nomeVeiculo como navio.
            query = '''
                SELECT TOP 1
                    ddg.numeroDi,
                    ddg.situacaoDi,
                    ddg.dataHoraSituacaoDi,
                    diTransp.nomeVeiculo,
                    diTransp.codigoViaTransporte,
                    diTransp.nomeTransportador,
                    diTransp.nomeVeiculo AS nomeNavio,
                    diFrete.valorTotalDolares AS frete_valorTotalDolares,
                    diFrete.totalReais AS frete_totalReais,
                    diSeguro.valorTotalDolares AS seguro_valorTotalDolares,
                    diSeguro.valorTotalReais AS seguro_valorTotalReais,
                    DVME.totalDolares AS valorMercadoriaEmbarque_totalDolares,
                    DVME.totalReais AS valorMercadoriaEmbarque_totalReais
                FROM Serpro.dbo.Di_Dados_Gerais ddg
                INNER JOIN Serpro.dbo.Di_Root_Declaracao_Importacao diRoot ON ddg.dadosGeraisId = diRoot.dadosGeraisId
                LEFT JOIN Serpro.dbo.Di_Transporte diTransp ON diRoot.transporteId = diTransp.transporteId
                LEFT JOIN Serpro.dbo.Di_Frete diFrete ON diRoot.dadosDiId = diFrete.freteId
                LEFT JOIN Serpro.dbo.Di_Seguro diSeguro ON diRoot.dadosDiId = diSeguro.seguroId
                LEFT JOIN Serpro.dbo.Di_Valor_Mercadoria_Embarque DVME 
                    ON diRoot.valorMercadoriaEmbarqueId = DVME.valorMercadoriaEmbarqueId
                WHERE ddg.numeroDi = ? OR ddg.numeroDi = ?
                ORDER BY ddg.dataHoraSituacaoDi DESC
            '''
            
            result = sql_adapter.execute_query(query, 'Serpro', [numero_di, numero_di_normalizado], notificar_erro=False)
            
            # ✅ CORREÇÃO: Tratar erro de tabela não encontrada silenciosamente
            if not result.get('success'):
                error_msg = result.get('error', '')
                if 'Invalid object name' in error_msg and 'Di_Dados_Embarque' in error_msg:
                    # Tabela não existe, mas como é LEFT JOIN, não deve causar erro fatal
                    # Apenas logar como debug
                    logger.debug(f'⚠️ Tabela Di_Dados_Embarque não encontrada para DI {numero_di}. Continuando sem dados de embarque...')
                else:
                    # Se for outro erro, logar como debug (não warning para não poluir logs)
                    logger.debug(f'⚠️ Erro ao buscar DI {numero_di} no SQL Server: {error_msg}')
                return None
            
            if result.get('data'):
                rows = result['data']
                if rows and len(rows) > 0:
                    row = rows[0]
                    
                    # Buscar pagamentos/impostos da DI do SQL Server
                    pagamentos_di = []
                    try:
                        dados_di_id = None
                        # Buscar dadosDiId para buscar pagamentos
                        query_di_id = '''
                            SELECT TOP 1 diRoot.dadosDiId
                            FROM Serpro.dbo.Di_Dados_Gerais ddg
                            INNER JOIN Serpro.dbo.Di_Root_Declaracao_Importacao diRoot ON ddg.dadosGeraisId = diRoot.dadosGeraisId
                            WHERE ddg.numeroDi = ? OR ddg.numeroDi = ?
                            ORDER BY ddg.dataHoraSituacaoDi DESC
                        '''
                        result_di_id = sql_adapter.execute_query(query_di_id, 'Serpro', [numero_di, numero_di_normalizado], notificar_erro=False)
                        if result_di_id.get('success') and result_di_id.get('data') and len(result_di_id['data']) > 0:
                            dados_di_id = result_di_id['data'][0].get('dadosDiId')
                            
                            if dados_di_id:
                                # Buscar pagamentos
                                query_pagamentos = '''
                                    SELECT 
                                        dp.codigoReceita,
                                        dp.numeroRetificacao,
                                        dp.valorTotal,
                                        dp.dataPagamento,
                                        dp.dataHoraPagamento,
                                        dpcr.descricao_receita
                                    FROM Serpro.dbo.Di_Pagamento dp
                                    LEFT JOIN Serpro.dbo.Di_pagamentos_cod_receitas dpcr 
                                        ON dpcr.cod_receita = dp.codigoReceita
                                    WHERE dp.rootDiId = ?
                                '''
                                result_pag = sql_adapter.execute_query(query_pagamentos, 'Serpro', [dados_di_id], notificar_erro=False)
                                if result_pag.get('success') and result_pag.get('data'):
                                    for pag_row in result_pag['data']:
                                        pagamentos_di.append({
                                            'codigoReceita': str(pag_row.get('codigoReceita') or ''),
                                            'valorTotal': str(pag_row.get('valorTotal') or '0'),
                                            'dataPagamento': pag_row.get('dataPagamento') or pag_row.get('dataHoraPagamento'),
                                            'descricao_receita': pag_row.get('descricao_receita') or ''
                                        })
                    except Exception as e:
                        logger.debug(f'Erro ao buscar pagamentos da DI {numero_di_normalizado}: {e}')
                    
                    # Montar estrutura similar à API para compatibilidade
                    dados_di = {
                        'numeroDi': row.get('numeroDi'),
                        'situacaoDi': row.get('situacaoDi'),
                        'navio': {
                            'nome_veiculo': row.get('nomeVeiculo') or row.get('nomeNavio') or ''
                        },
                        'transporte': {
                            'nomeVeiculo': row.get('nomeVeiculo') or '',
                            'nomeNavio': row.get('nomeNavio') or '',
                            'codigoViaTransporte': row.get('codigoViaTransporte') or '',
                            'nomeTransportador': row.get('nomeTransportador') or ''
                        },
                        'frete': {
                            'valorTotalDolares': str(row.get('frete_valorTotalDolares') or '0'),
                            'totalReais': str(row.get('frete_totalReais') or '0')
                        },
                        'seguro': {
                            'valorTotalDolares': str(row.get('seguro_valorTotalDolares') or '0'),
                            'valorTotalReais': str(row.get('seguro_valorTotalReais') or '0')
                        },
                        'valorMercadoriaEmbarque': {
                            'totalDolares': str(row.get('valorMercadoriaEmbarque_totalDolares') or '0'),
                            'totalReais': str(row.get('valorMercadoriaEmbarque_totalReais') or '0')
                        },
                        'pagamentos': pagamentos_di  # ✅ Pagamentos do SQL Server
                    }
                    
                    nome_veiculo = dados_di['navio']['nome_veiculo'] or dados_di['transporte']['nomeVeiculo'] or ''
                    logger.info(f'✅ Dados da DI {numero_di_normalizado} encontrados no SQL Server (nomeVeiculo: {nome_veiculo}, {len(pagamentos_di)} pagamentos)')
                    return dados_di
            
            return None
            
        except Exception as e:
            logger.warning(f'⚠️ Erro ao buscar DI {numero_di_normalizado} no SQL Server: {e}')
            return None
    
    def _extrair_dados_processo(self, processo: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Extrai dados completos de um processo para o relatório.
        
        Busca dados da DI seguindo prioridade: SQL Server → API Integra Comex (bilhetada).
        O nome do navio está disponível no SQL Server na tabela Di_Transporte (campo nomeVeiculo).
        
        Args:
            processo: Dict com processo_referencia, numero_di, etc.
        
        Returns:
            Dict com dados da linha do relatório ou None se erro
        """
        try:
            numero_di = processo.get('numero_di')
            if not numero_di:
                logger.warning(f'⚠️ Processo {processo.get("processo_referencia")} não tem número de DI')
                return None
            
            # Normalizar número da DI
            numero_di_normalizado = numero_di.replace('/', '').replace('-', '').replace(' ', '').replace('.', '')
            
            # ✅ PRIORIDADE 1: Buscar dados da DI do SQL Server primeiro (sem custo)
            dados_di_sql = self._buscar_di_sql_server(numero_di, numero_di_normalizado)
            
            # ✅ PRIORIDADE 2: Se não encontrou no SQL Server, buscar via API (bilhetada)
            if not dados_di_sql:
                logger.info(f'📡 DI {numero_di_normalizado} não encontrada no SQL Server, buscando via API...')
                from services.di_pdf_service import DiPdfService
                
                di_service = DiPdfService()
                status, dados_di_api = di_service.consultar_di_integracomex(numero_di_normalizado)
                
                if status == 200 and isinstance(dados_di_api, dict):
                    dados_di_sql = dados_di_api
                else:
                    logger.warning(f'⚠️ Erro ao buscar DI {numero_di_normalizado} via API: status={status}')
                    return None
            else:
                logger.info(f'✅ DI {numero_di_normalizado} encontrada no SQL Server')
            
            # Extrair campos da DI
            dados_linha = self._mapear_dados_di_para_linha(dados_di_sql, processo)
            
            return dados_linha
            
        except Exception as e:
            logger.error(f'❌ Erro ao extrair dados do processo {processo.get("processo_referencia")}: {e}', exc_info=True)
            return None
    
    def _mapear_dados_di_para_linha(
        self,
        dados_di: Dict[str, Any],
        processo: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Mapeia dados da DI para formato da linha do relatório.
        
        Args:
            dados_di: Dados completos da DI da API
            processo: Dados do processo
        
        Returns:
            Dict com dados da linha
        """
        # Extrair seções da DI
        conhecimento_carga = dados_di.get('conhecimentoCarga', {}) or {}
        frete = dados_di.get('frete', {}) or {}
        transporte = dados_di.get('transporte', {}) or {}
        valor_embarque = dados_di.get('valorMercadoriaEmbarque', {}) or {}
        pagamentos = dados_di.get('pagamentos', []) or []
        icms = dados_di.get('icms', []) or []
        adicoes = dados_di.get('adiccoes', []) or []
        
        # País de Origem
        pais_procedencia = dados_di.get('paisProcedencia', '')
        pais_origem = PAISES.get(pais_procedencia, pais_procedencia)
        
        # Porto Origem
        porto_origem = conhecimento_carga.get('localEmbarque', '')
        
        # Data do BL
        data_bl = conhecimento_carga.get('dataEmbarque', '')
        
        # Tipo de transporte
        codigo_via = transporte.get('codigoViaTransporte', '')
        tipo_transporte = TIPOS_TRANSPORTE.get(str(codigo_via), 'NAVIO')
        
        # Nome Navio
        # ✅ CORREÇÃO: Buscar nomeVeiculo do SQL Server (Di_Transporte) ou da API
        # Prioridade: navio (SQL Server) → transporte.nomeVeiculo (API) → transporte.nomeNavio (API)
        nome_navio = dados_di.get('navio', {}).get('nome_veiculo', '') if isinstance(dados_di.get('navio'), dict) else ''
        if not nome_navio:
            nome_navio = transporte.get('nomeVeiculo', '')
        if not nome_navio:
            nome_navio = transporte.get('nomeNavio', '')
        
        # Custo USD
        custo_usd = float(valor_embarque.get('totalDolares', 0) or 0)
        
        # Frete USD
        frete_usd = float(frete.get('valorTotalDolares', 0) or 0)
        
        # Despesas (10% de Custo + Frete)
        despesas = (custo_usd + frete_usd) * 0.10
        
        # Lucros (10% de Custo + Frete + Despesas)
        lucros = (custo_usd + frete_usd + despesas) * 0.10
        
        # Impostos da DI USD (somar todos os pagamentos + ICMS)
        # ⚠️ IMPORTANTE: Os pagamentos vêm em BRL, precisamos converter para USD
        # Usar PTAX do dia do pagamento ou data de registro da DI
        impostos_total_usd = 0.0
        
        # Data de referência para conversão (usar data de registro da DI ou data do pagamento)
        data_referencia = processo.get('data_registro')
        if data_referencia:
            if isinstance(data_referencia, str):
                # Parsear data
                try:
                    if 'T' in data_referencia:
                        data_referencia_dt = datetime.fromisoformat(data_referencia.split('.')[0])
                    else:
                        data_referencia_dt = datetime.strptime(data_referencia, '%Y-%m-%d')
                except:
                    data_referencia_dt = datetime.now()
            else:
                data_referencia_dt = data_referencia
        else:
            data_referencia_dt = datetime.now()
        
        # Buscar PTAX para conversão
        try:
            from utils.ptax_bcb import obter_ptax_dolar
            data_ptax_str = data_referencia_dt.strftime('%m-%d-%Y')
            ptax_result = obter_ptax_dolar(data_ptax_str)
            if ptax_result and ptax_result.get('sucesso'):
                cotacao_ptax = ptax_result.get('cotacao_media', 1.0)
            else:
                logger.warning(f'⚠️ PTAX não encontrada para {data_ptax_str}, usando 1.0')
                cotacao_ptax = 1.0
        except Exception as e:
            logger.warning(f'⚠️ Erro ao buscar PTAX: {e}, usando 1.0')
            cotacao_ptax = 1.0
        
        # Somar pagamentos (converter BRL para USD)
        for pagamento in pagamentos:
            codigo_receita = str(pagamento.get('codigoReceita', ''))
            valor_total_brl = float(pagamento.get('valorTotal', 0) or 0)
            
            # Converter BRL para USD usando PTAX
            valor_total_usd = valor_total_brl / cotacao_ptax if cotacao_ptax > 0 else 0
            impostos_total_usd += valor_total_usd
        
        # Somar ICMS (se houver) - também em BRL, converter para USD
        for icms_item in icms:
            valor_icms_brl = float(icms_item.get('valor', 0) or 0)
            valor_icms_usd = valor_icms_brl / cotacao_ptax if cotacao_ptax > 0 else 0
            impostos_total_usd += valor_icms_usd
        
        # Mercadoria (pegar primeira adição ou concatenar todas)
        mercadoria = ''
        if adicoes:
            primeira_adicao = adicoes[0]
            mercadoria_obj = primeira_adicao.get('mercadoria', {}) or {}
            mercadoria = mercadoria_obj.get('descricaoMercadoria', '')
        
        # Cidade de Destino (buscar do CE ou da DI)
        cidade_destino = ''
        try:
            # Tentar buscar do processo consolidado (CE)
            processo_ref = processo.get('processo_referencia')
            if processo_ref:
                from services.sql_server_processo_schema import buscar_processo_consolidado_sql_server
                processo_consolidado = buscar_processo_consolidado_sql_server(processo_ref)
                if processo_consolidado and processo_consolidado.get('ce'):
                    porto_destino = processo_consolidado['ce'].get('porto_destino', '')
                    # Mapeamento básico de códigos de porto para cidades
                    # TODO: Expandir mapeamento ou buscar de tabela/API
                    porto_cidade_map = {
                        'BRIGI': 'Itaguai',
                        'BRRIO': 'Rio de Janeiro',
                        'BRSPS': 'São Paulo',
                        'BRSSZ': 'Santos',
                        # Adicionar mais conforme necessário
                    }
                    cidade_destino = porto_cidade_map.get(porto_destino, porto_destino)
        except Exception as e:
            logger.debug(f'Erro ao buscar cidade de destino: {e}')
        
        # DI
        numero_di = dados_di.get('numeroDi', '') or processo.get('numero_di', '')
        
        # OBS (processo_referencia)
        obs = processo.get('processo_referencia', '')
        
        return {
            'pais_origem': pais_origem,
            'porto_origem': porto_origem,
            'pais_origem_2': pais_origem,  # Duplicado na planilha
            'cidade_destino': cidade_destino,
            'data_bl': data_bl,
            'tipo_transporte': tipo_transporte,
            'mercadoria': mercadoria,
            'nome_navio': nome_navio,
            'custo_usd': custo_usd,
            'frete_usd': frete_usd,
            'despesas': despesas,
            'lucros': lucros,
            'impostos_usd': impostos_total_usd,
            'di': numero_di,
            'obs': obs
        }
    
    def _gerar_excel(
        self,
        dados: List[Dict[str, Any]],
        mes: int,
        ano: int,
        categoria: Optional[str] = None
    ) -> Path:
        """
        Gera arquivo Excel no formato da planilha de averbações.
        
        Args:
            dados: Lista de dicts com dados das linhas
            mes: Mês
            ano: Ano
            categoria: Categoria (opcional)
        
        Returns:
            Path do arquivo gerado
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'IMPORTAÇÃO'
            
            # Cabeçalho (linhas 1-4)
            ws['A1'] = 'Modelo Averbação Via Planilha'
            ws['A2'] = 'Nome do Segurado'
            ws['C2'] = 'BANDEMAR COMERCIO IMPORTACAO E EXPORTACAO'
            ws['A3'] = 'CNPJ'
            ws['C3'] = '08.641.586/0002-66 e 08.641.586/0004-28'
            ws['A4'] = 'Nº Apólice'
            
            # Cabeçalhos das colunas (linha 5)
            headers = [
                'Averbação provisória',
                'País de Origem',
                'Porto Origem',
                'País Origem',
                'Cidade de  Destino',
                'Data do BL',
                'Tipo de transporte',
                'Mercadoria',
                'Nome Navio',
                'Custo USD',
                'Frete USD',
                'Despesas',
                'Lucros',
                'Impostos da DI USD',
                'DI',
                'OBS'
            ]
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Dados (linha 6 em diante)
            for row_idx, linha in enumerate(dados, start=6):
                ws.cell(row=row_idx, column=1).value = ''  # Averbação provisória (vazio)
                ws.cell(row=row_idx, column=2).value = linha.get('pais_origem', '')
                ws.cell(row=row_idx, column=3).value = linha.get('porto_origem', '')
                ws.cell(row=row_idx, column=4).value = linha.get('pais_origem_2', '')
                ws.cell(row=row_idx, column=5).value = linha.get('cidade_destino', '')
                ws.cell(row=row_idx, column=6).value = linha.get('data_bl', '')
                ws.cell(row=row_idx, column=7).value = linha.get('tipo_transporte', '')
                ws.cell(row=row_idx, column=8).value = linha.get('mercadoria', '')
                ws.cell(row=row_idx, column=9).value = linha.get('nome_navio', '')
                ws.cell(row=row_idx, column=10).value = linha.get('custo_usd', 0)
                ws.cell(row=row_idx, column=11).value = linha.get('frete_usd', 0)
                ws.cell(row=row_idx, column=12).value = linha.get('despesas', 0)
                ws.cell(row=row_idx, column=13).value = linha.get('lucros', 0)
                ws.cell(row=row_idx, column=14).value = linha.get('impostos_usd', 0)
                ws.cell(row=row_idx, column=15).value = linha.get('di', '')
                ws.cell(row=row_idx, column=16).value = linha.get('obs', '')
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 40
            ws.column_dimensions['I'].width = 25
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 12
            ws.column_dimensions['M'].width = 12
            ws.column_dimensions['N'].width = 18
            ws.column_dimensions['O'].width = 15
            ws.column_dimensions['P'].width = 15
            
            # Salvar arquivo
            categoria_str = f'_{categoria}' if categoria else ''
            nome_arquivo = f'Relatorio_Averbacoes_{ano:04d}_{mes:02d}{categoria_str}.xlsx'
            caminho_arquivo = self.downloads_dir / nome_arquivo
            
            wb.save(caminho_arquivo)
            logger.info(f'✅ Excel gerado: {caminho_arquivo}')
            
            return caminho_arquivo
            
        except Exception as e:
            logger.error(f'❌ Erro ao gerar Excel: {e}', exc_info=True)
            raise
